# -*- coding: utf-8 -*-
"""Helpers for building training subsets from ``raw_data.xlsx``.

The source spreadsheet contains hourly traffic counts split across multiple
worksheets (one per month and weekday/weekend group).  Each data row stores the
route metadata followed by 24 hourly measurements.  For lightweight regression
experiments we treat the first 12 hourly values as input features and the
remaining 12 as the target that should be forecast.

This module exposes a small utility that extracts a handful of rows from
selected worksheets and converts them into dictionaries that are directly
consumable by :func:`stnet.api.run.train`.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Iterable, Iterator, List, Sequence, Tuple

import torch

try:
    import openpyxl  # type: ignore
except Exception as exc:  # pragma: no cover - dependency is optional
    raise RuntimeError(
        "openpyxl is required to materialise raw-data subsets"
    ) from exc


FeatureKey = Tuple[float, ...]


@dataclass(frozen=True)
class SubsetSpec:
    """Selection instructions for :func:`build_excel_subsets`.

    Attributes
    ----------
    sheet:
        Name of the worksheet to sample from.
    rows:
        Iterable of zero-based row indices **within the sheet's data region**.
        The header row is skipped automatically, so ``rows=range(0, 5)`` selects
        the first five data rows beneath the header.
    name:
        Optional display name.  When omitted, ``sheet`` is used instead.
    """

    sheet: str
    rows: Iterable[int]
    name: str | None = None


@dataclass(frozen=True)
class Subset:
    """Represents a small Excel-backed training subset."""

    name: str
    sheet: str
    row_indices: Tuple[int, ...]
    samples: Dict[FeatureKey, torch.Tensor]

    def __iter__(self) -> Iterator[Tuple[FeatureKey, torch.Tensor]]:
        return iter(self.samples.items())


def _row_to_tensors(row: Sequence[object]) -> Tuple[FeatureKey, torch.Tensor]:
    """Convert a worksheet row into a feature/label pair.

    Parameters
    ----------
    row:
        Sequence of cell values (already stripped of header metadata).

    Returns
    -------
    Tuple[FeatureKey, torch.Tensor]
        The feature tuple (12 hourly measurements) and a ``torch.Tensor`` label
        containing the remaining 12 hourly measurements.
    """

    if len(row) < 24:
        raise ValueError("expected at least 24 hourly values per row")
    # First three entries correspond to line/segment/direction metadata.
    payload = row[3:]
    if len(payload) < 24:
        raise ValueError("row is missing hourly payload columns")
    try:
        floats = [float(v) for v in payload]
    except (TypeError, ValueError) as exc:
        raise ValueError(f"non-numeric value encountered in row: {row!r}") from exc
    features = tuple(floats[:12])
    targets = torch.tensor(floats[12:], dtype=torch.float64)
    return (features, targets)


def build_excel_subsets(path: str, specs: Sequence[SubsetSpec]) -> List[Subset]:
    """Materialise a list of small subsets from ``raw_data.xlsx``.

    Parameters
    ----------
    path:
        Filesystem path to the Excel workbook.
    specs:
        Subset selection descriptors.

    Returns
    -------
    List[Subset]
        The list of extracted subsets ready for consumption by ``train`` or
        ``predict``.
    """

    if not specs:
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    subsets: List[Subset] = []
    for spec in specs:
        if spec.sheet not in wb.sheetnames:
            raise KeyError(f"worksheet '{spec.sheet}' not found in {path!r}")
        ws = wb[spec.sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        data_rows = rows[1:]
        samples: Dict[FeatureKey, torch.Tensor] = {}
        resolved_indices: List[int] = []
        for row_idx in spec.rows:
            if row_idx < 0 or row_idx >= len(data_rows):
                raise IndexError(
                    f"row index {row_idx} out of range for sheet '{spec.sheet}'"
                )
            resolved_indices.append(int(row_idx))
            row = data_rows[row_idx]
            if row is None:
                continue
            feat_key, targets = _row_to_tensors(row)
            samples[feat_key] = targets
        subset_name = spec.name if spec.name else spec.sheet
        subsets.append(
            Subset(
                name=subset_name,
                sheet=spec.sheet,
                row_indices=tuple(resolved_indices),
                samples=samples,
            )
        )
    return subsets


__all__ = ["Subset", "SubsetSpec", "build_excel_subsets"]

