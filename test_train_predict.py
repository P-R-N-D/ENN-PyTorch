#!/usr/bin/env python3
"""Train and predict using the bundled raw_data.xlsx workbook.

This script mirrors the workflow demonstrated in the project notebook while
keeping runtime short so it can be used as an integration smoke test. It loads
`raw_data.xlsx`, materializes training tensors, runs a single epoch of
`stnet.workflow.train`, and validates inference through `stnet.workflow.predict`.

All temporary artifacts created by the workflow helpers (memory-mapped
intermediates, checkpoints, distributed rendezvous files, etc.) are removed
before the script exits.
"""
from __future__ import annotations

import argparse
import math
import os
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import numpy as np
import polars as pl
import torch
from openpyxl import load_workbook

from stnet.workflow import Config, new_model, predict, train

HOURS = [f"{h:02d}시" for h in range(24)]
DAY2ID = {"평일": 0, "주말·공휴일": 1}
DIR2ID = {"상행": 0, "하행": 1}


def _is_hour_col(name: str) -> bool:
    text = str(name)
    return len(text) == 3 and text.endswith("시") and text[:2].isdigit()


def _schema_names(df: pl.DataFrame | pl.LazyFrame) -> List[str]:
    if isinstance(df, pl.DataFrame):
        return list(df.columns)
    if isinstance(df, pl.LazyFrame):
        if hasattr(df, "schema"):
            schema = df.schema
            if hasattr(schema, "names"):
                return list(schema.names())
            if isinstance(schema, dict):
                return list(schema.keys())
        if hasattr(df, "collect_schema"):
            return list(df.collect_schema().names())
    raise TypeError("df must be a Polars DataFrame or LazyFrame")


def _parse_sheet_name(name: str) -> Tuple[int | None, str]:
    digits = "".join(ch for ch in name if ch.isdigit())
    month = int(digits) if digits else None
    daytype = "평일" if "평일" in name else "주말·공휴일"
    return (month, daytype)


def _cleanse_sheet(
    frame: pl.DataFrame | pl.LazyFrame, *, month: int, daytype: str
) -> pl.DataFrame:
    hour_cols = [c for c in HOURS if c in _schema_names(frame)]
    lazy_frame = frame.lazy() if isinstance(frame, pl.DataFrame) else frame
    return (
        lazy_frame.with_columns(
            pl.col("구간").cast(pl.String),
            pl.col("방향").cast(pl.String),
        )
        .filter(
            pl.col("구간").is_not_null()
            & pl.col("구간").str.contains("-", literal=True)
        )
        .with_columns([pl.col(col).cast(pl.Float32, strict=False) for col in hour_cols])
        .with_columns(
            pl.col("구간").str.replace_all(r"\s+", "").alias("구간"),
            pl.col("방향").str.strip_chars().alias("방향"),
            pl.col("구간").str.split("-").list.sort().list.join("-").alias("구간ID"),
        )
        .unpivot(
            index=["구간", "구간ID", "방향"],
            on=hour_cols,
            variable_name="시간문자",
            value_name="속도",
        )
        .with_columns(
            pl.lit(month).alias("월"),
            pl.lit(DAY2ID[daytype]).alias("요일타입_id"),
            pl.when(pl.col("방향") == "상행")
            .then(0)
            .when(pl.col("방향") == "하행")
            .then(1)
            .otherwise(None)
            .cast(pl.Int8, strict=False)
            .alias("방향_id"),
            pl.col("시간문자")
            .str.replace_all("시", "")
            .cast(pl.Int32, strict=False)
            .alias("시간"),
            pl.col("속도").cast(pl.Float32, strict=False),
        )
        .select(["월", "요일타입_id", "방향_id", "구간ID", "시간", "속도"])
        .filter(pl.col("방향_id").is_not_null())
        .group_by(["월", "요일타입_id", "방향_id", "구간ID", "시간"])
        .agg(pl.col("속도").mean().alias("속도"))
        .collect()
    )


def _nanmean_grid(grid: np.ndarray) -> np.ndarray:
    with np.errstate(invalid="ignore"):
        col_mean = np.nanmean(grid, axis=0, keepdims=True)
    col_mean = np.where(np.isnan(col_mean), 0.0, col_mean)
    return np.where(np.isnan(grid), col_mean, grid)


def _load_raw_tensor(path: Path) -> Tuple[torch.Tensor, torch.Tensor, List[str]]:
    if not path.exists():
        raise FileNotFoundError(f"raw workbook not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = workbook.sheetnames
        raw = {
            name: pl.read_excel(path, sheet_name=name).lazy() for name in sheet_names
        }
    finally:
        workbook.close()

    long_frames: List[pl.DataFrame] = []
    for name in sheet_names:
        month, daytype = _parse_sheet_name(name)
        if month is None:
            raise ValueError(f"sheet name is missing month digits: {name}")
        long_frames.append(_cleanse_sheet(raw[name], month=month, daytype=daytype))

    combined = pl.concat(long_frames, how="vertical", rechunk=True)
    segments = (
        combined.select("구간ID")
        .unique()
        .sort("구간ID")
        .get_column("구간ID")
        .to_list()
    )
    seg_to_index = {seg: idx for idx, seg in enumerate(segments)}
    hours = 24
    feat_to_tensor: Dict[Tuple[int, int, int], torch.Tensor] = {}

    for (month, day_id, dir_id), group in combined.group_by(
        ["월", "요일타입_id", "방향_id"]
    ):
        grid = np.full((hours, len(segments), 1), np.nan, dtype=np.float32)
        for hour, seg, value in group.select(["시간", "구간ID", "속도"]).iter_rows():
            seg_index = seg_to_index[seg]
            if value is not None and not math.isnan(value):
                grid[int(hour), seg_index, 0] = float(value)
        grid = _nanmean_grid(grid)
        feat_to_tensor[(int(month), int(day_id), int(dir_id))] = torch.from_numpy(grid)

    ordered_keys = sorted(feat_to_tensor.keys())
    features = torch.tensor(ordered_keys, dtype=torch.float32)
    labels = torch.stack([feat_to_tensor[key] for key in ordered_keys], dim=0)
    return (features, labels, segments)


def _build_small_config(batch_size: int) -> Config:
    return Config(
        device="cpu",
        microbatch=max(1, min(batch_size, 8)),
        dropout=0.0,
        depth=16,
        heads=2,
        spatial_depth=2,
        temporal_depth=2,
        mlp_ratio=2.0,
        drop_path=0.0,
        spatial_latent_tokens=16,
        temporal_latent_tokens=16,
        use_linear_branch=False,
        use_compilation=False,
    )


def _predict_inputs(features: torch.Tensor, out_shape: Sequence[int]) -> Dict[Tuple[int, ...], torch.Tensor | None]:
    dummy = torch.zeros(out_shape, dtype=torch.float32)
    result: Dict[Tuple[int, ...], torch.Tensor | None] = {}
    for row in features:
        key = tuple(int(v.item()) for v in row)
        result[key] = None if dummy.numel() else dummy
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path(os.environ.get("STNET_RAW_XLSX", "raw_data.xlsx")),
        help="Path to raw_data.xlsx (defaults to repository copy).",
    )
    parser.add_argument("--epochs", type=int, default=1)
    parser.add_argument("--batch-size", type=int, default=8)
    parser.add_argument("--val-frac", type=float, default=0.2)
    parser.add_argument("--base-lr", type=float, default=3e-3)
    parser.add_argument("--weight-decay", type=float, default=1e-4)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    features, labels, segments = _load_raw_tensor(args.xlsx)
    print(f"Loaded {features.shape[0]} samples across {len(segments)} segments.")
    batch_size = max(1, min(args.batch_size, features.shape[0]))
    config = _build_small_config(batch_size)
    model = new_model(int(features.shape[1]), tuple(labels.shape[1:]), config)

    trained = train(
        model,
        {"X": features, "Y": labels},
        epochs=max(1, args.epochs),
        batch_size=batch_size,
        val_frac=max(0.0, min(args.val_frac, 0.5)),
        base_lr=float(args.base_lr),
        weight_decay=float(args.weight_decay),
        max_nodes=1,
        prefetch_factor=1,
    )

    infer_inputs = _predict_inputs(features, labels.shape[1:])
    predictions = predict(
        trained,
        infer_inputs,
        batch_size=batch_size,
        prefetch_factor=1,
    )
    print(f"Generated {len(predictions)} prediction grids.")
    sample_key = next(iter(predictions))
    sample_grid = predictions[sample_key]
    print(f"Sample key: {sample_key}, prediction tensor shape: {tuple(sample_grid.shape)}")


if __name__ == "__main__":
    torch.manual_seed(0)
    main()
